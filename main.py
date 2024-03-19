import time
from openpyxl import load_workbook
from pynput.keyboard import Key, Controller
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys

filepath = "C:/Users/Raghav/OneDrive/Desktop/contact_list.xlsx"
wb = load_workbook(filepath)
sheet = wb.active
frequency = 7997
# open chrome
driver = webdriver.Chrome('C:\chromedriver\chromedriver')
# open whatsapp homepage for scanning
driver.get("https://web.whatsapp.com")
# wait for some time(in sec)
time.sleep(120)
# open new window
driver.execute_script("window.open('');")
keyboard = Controller()


def shortcut():
    driver.find_element_by_xpath('//*[@title="Type a message"]').send_keys(Keys.SHIFT, Keys.ENTER)


def newLine():
    driver.find_element_by_xpath('//*[@title="Type a message"]').send_keys(Keys.SHIFT, Keys.ENTER)
    driver.find_element_by_xpath('//*[@title="Type a message"]').send_keys(Keys.SHIFT, Keys.ENTER)


i = 0
l = 0
while i <= frequency:
    while i <= frequency:
        b3 = sheet.cell(row=i + 1, column=1)
        print("Sr. no." + (str(i + 1)))
        # Switch to the new window
        driver.switch_to.window(driver.window_handles[1])
        # open the link
        url = "https://web.whatsapp.com/send?phone=+91" + str(b3.value) + "&text=" + "&app_absent=0"
        driver.get(url)
        time.sleep(25)
        # attaching image
        try:
            driver.find_element_by_xpath('//div[@title = "Attach"]').click()
        except NoSuchElementException:
            print(str(b3.value) + " NOT ON WHATSAPP")
            l += 1
            i += 1
            break
        image_box = driver.find_element_by_xpath('//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"]')
        # locating image
        image_box.send_keys('C:/Users/Raghav/OneDrive/Desktop/Brochure.jpg')
        time.sleep(2)
        # pressing send button
        driver.find_element_by_xpath('//span[@data-icon="send"]').click()
        # wait for some time
        time.sleep(1)
        # typing message
        message = driver.find_element_by_xpath('//*[@title="Type a message"]')
        message.send_keys("DEAR PARENTS -")
        shortcut()
        message.send_keys(" Unlock Your Child's Potential! ")
        shortcut()
        message.send_keys("Discover the Best Classes for Your Child!")
        shortcut()
        message.send_keys(" Professional Guidance")
        shortcut()
        message.send_keys(" Personal Attention")
        shortcut()
        message.send_keys(
            "We offer specialized classes to nurture your child's talents and address areas for improvement. Our goal is to shape well-rounded individuals.")
        shortcut()
        shortcut()
        message.send_keys(" Creative Arts ")
        shortcut()
        message.send_keys("(Sketching, Drawing,Painting)")
        shortcut()
        message.send_keys(" Academic Mastery")
        shortcut()
        message.send_keys("(Tuitions,Phonic reading, Creative writing )")
        shortcut()
        message.send_keys(" Basketball")
        shortcut()
        message.send_keys(" Musical Instruments")
        shortcut()
        message.send_keys("(Guitar,Synthesizer,Dholak,Tabla)")
        shortcut()
        message.send_keys(" Dance")
        shortcut()
        message.send_keys("(Western, folk,Bollywood)")
        shortcut()
        message.send_keys(" Skating")
        shortcut()
        message.send_keys(" UCMAS")
        shortcut()
        message.send_keys(" Handwriting Improvement")
        shortcut()
        message.send_keys(" Calligraphy")
        shortcut()
        message.send_keys(" Chess")
        shortcut()
        shortcut()
        message.send_keys("BANIPARK: D-92(Z) MeeraMarg Banipark, Jaipur")
        shortcut()
        shortcut()
        message.send_keys(
            "Vidhyadhar Nagar: 144, Opp. MGPS, Shankar Colony Mall Road, Vidhyadhar Nagar, Jaipur, Rajasthan 302039")
        shortcut()
        message.send_keys("https://bit.ly/3Q62Emg")
        shortcut()
        shortcut()
        message.send_keys("Find us on Instagram -> www.instagram.com/moderngurukulacademy")
        shortcut()
        shortcut()
        message.send_keys("For More Details Contact: 9784597275 , 9680836271")

        # press enter
        keyboard.press(Key.enter)
        keyboard.release(Key.enter)
        print("Sent Messages = " + str((i + 1) - l))
        i += 1
        # wait for some time
        time.sleep(2)

# Close the browser
driver.quit()
print("You have circulated through all the numbers.")
print("https://www.youtube.com/watch?v=dQw4w9WgXcQ")
